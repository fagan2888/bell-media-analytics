from openpyxl import Workbook
from time import strftime, localtime
import re, auth

class harToExcel:

	def populate_sheet(self, ws, dic):
		i = 1
		for key in sorted(dic.keys()):
			c = ws.cell(row = i, column = 1)
			c.value = key
			c = ws.cell(row = i, column = 2)
			c.value = dic[key]
			i += 1

	def create_dictionaries(self, unparsed_har_file):
		# this function performs dark parsing magic
		f = open(unparsed_har_file)
		har_file = f.read()

		clump_of_data = re.findall('"request"\:\{"queryString"\:\[\{"name"\:"AQB"[^\]]*', har_file)

		list_of_dictionaries = tmp = []

		for i in range(0, len(clump_of_data)):
			tmp = re.findall('\{[^\}]*\}', clump_of_data[i])
			list_of_dictionaries.append({})

			for j in range(0, len(tmp)):
				l = re.findall('"[^"]*"', tmp[j])
				l = map(lambda x: x.replace("\"",""), l)

				l.remove("name")
				l.remove("value")

				if "queryString" in l:
					l.remove("queryString")
				
				if len(l) == 2:
					list_of_dictionaries[i].update({l[0]:l[1]})
				
		return list_of_dictionaries

	def create_document(self, har_location, save_location):
		try:
			list_of_dictionaries = self.create_dictionaries(har_location)
			wb = Workbook()
			ws = wb.get_active_sheet()
			ws.title = "Sheet 1"
			self.populate_sheet(ws, list_of_dictionaries[0])

			for i in range(1,len(list_of_dictionaries)):
				ws = wb.create_sheet()
				self.populate_sheet(ws, list_of_dictionaries[i])
				ws.title = "Sheet " + str(i+1)

			wb.save(save_location)
			print "HAR to XLSX conversion complete!"
		except:
			print "HAR to XLSX conversion FAILED! (probably because of Fiddler)"